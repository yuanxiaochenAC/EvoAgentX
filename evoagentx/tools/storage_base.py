import json
import pickle
import csv
import yaml
import xml.etree.ElementTree as ET
import os
from typing import Dict, Any, List
from pathlib import Path
from abc import ABC, abstractmethod

# For handling various file types
try:
    import pymupdf
    PDF_AVAILABLE = True
except ImportError:
    PDF_AVAILABLE = False

try:
    from PIL import Image
    PILLOW_AVAILABLE = True
except ImportError:
    PILLOW_AVAILABLE = False

try:
    from openpyxl import Workbook, load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

from ..core.module import BaseModule
from ..core.logging import logger


class StorageBase(BaseModule, ABC):
    """
    Abstract base class for comprehensive storage operations supporting various file types.
    Provides unified interface for local and remote storage operations.
    """
    
    def __init__(self, base_path: str = ".", **kwargs):
        """
        Initialize the StorageBase with configuration options.
        
        Args:
            base_path (str): Base directory for storage operations (default: current directory)
            **kwargs: Additional keyword arguments for parent class initialization
        """
        super().__init__(**kwargs)
        self.base_path = base_path
        
        # File types that support append operations
        self.appendable_formats = {
            '.txt': self._append_text,
            '.json': self._append_json,
            '.csv': self._append_csv,
            '.yaml': self._append_yaml,
            '.yml': self._append_yaml,
            '.pickle': self._append_pickle,
            '.xlsx': self._append_excel
        }
        
        # Initialize storage-specific setup
        self._initialize_storage()
    
    @abstractmethod
    def _initialize_storage(self):
        """
        Initialize storage-specific setup. Override in subclasses for storage-specific initialization.
        """
        pass
    
    
    # Abstract methods that must be implemented by subclasses
    @abstractmethod
    def _read_raw(self, path: str, **kwargs) -> bytes:
        """Read raw file content - must be implemented by subclasses"""
        pass
    
    @abstractmethod
    def _write_raw(self, path: str, content: bytes, **kwargs) -> bool:
        """Write raw file content - must be implemented by subclasses"""
        pass
    
    @abstractmethod
    def _delete_raw(self, path: str) -> bool:
        """Delete file or directory - must be implemented by subclasses"""
        pass
    
    @abstractmethod
    def _list_raw(self, path: str = None, **kwargs) -> List[Dict[str, Any]]:
        """List files and directories - must be implemented by subclasses"""
        pass
    
    @abstractmethod
    def _exists_raw(self, path: str) -> bool:
        """Check if path exists - must be implemented by subclasses"""
        pass
    
    @abstractmethod
    def _create_directory_raw(self, path: str) -> bool:
        """Create directory - must be implemented by subclasses"""
        pass
    
    
    # ____________________ PATH TRANSLATION ____________________ #
    def translate_in(self, file_path: str) -> str:
        """
        Translate input file path by combining it with base_path.
        This method takes a user-provided path and converts it to the full system path.
        
        Args:
            file_path (str): User-provided file path (can be relative or absolute)
            
        Returns:
            str: Full system path combining base_path and file_path
        """
        # If the path is already absolute, return as is
        if os.path.isabs(file_path):
            return file_path
        
        # Always combine base_path with file_path to ensure working directory is respected
        # Check if this is a remote storage handler (like Supabase)
        if hasattr(self, 'bucket_name') and hasattr(self, 'supabase'):
            # For remote storage, treat base_path as a prefix within the bucket
            # Don't use os.path.join as it's designed for local filesystems
            if self.base_path.startswith('/'):
                # Remove leading slash and combine
                clean_base = self.base_path.lstrip('/')
                if clean_base:
                    return f"{clean_base}/{file_path}"
                else:
                    return file_path
            else:
                # Combine base_path and file_path with forward slash
                return f"{self.base_path}/{file_path}"
        else:
            # For local storage, use os.path.join for proper filesystem handling
            combined_path = os.path.join(self.base_path, file_path)
            normalized_path = os.path.normpath(combined_path)
            return normalized_path
    
    def translate_out(self, full_path: str) -> str:
        """
        Translate output full path by removing the base_path prefix.
        This method takes a full system path and converts it back to the user-relative path.
        
        Args:
            full_path (str): Full system path
            
        Returns:
            str: User-relative path with base_path removed
        """
        # If base_path is just "." or empty, return the full_path as is
        if self.base_path in [".", "", None]:
            return full_path
        
        # Check if this is a remote storage handler (like Supabase)
        if hasattr(self, 'bucket_name') and hasattr(self, 'supabase'):
            # For remote storage, handle path prefix removal
            if self.base_path.startswith('/'):
                clean_base = self.base_path.lstrip('/')
            else:
                clean_base = self.base_path
            
            if clean_base and full_path.startswith(f"{clean_base}/"):
                # Remove the base_path prefix
                relative_path = full_path[len(f"{clean_base}/"):]
                return relative_path
            elif clean_base and full_path == clean_base:
                # If the full_path is exactly the base_path, return empty string
                return ""
            else:
                # If the path doesn't start with base_path, return as is
                return full_path
        else:
            # For local storage, use os.path operations for proper filesystem handling
            # Convert both paths to absolute paths for comparison
            base_abs = os.path.abspath(self.base_path)
            full_abs = os.path.abspath(full_path)
            
            # Check if the full_path starts with base_path
            if full_abs.startswith(base_abs):
                # Remove the base_path prefix
                relative_path = full_abs[len(base_abs):]
                # Remove leading separator if present
                if relative_path.startswith(os.sep):
                    relative_path = relative_path[1:]
                return relative_path
            
            # If the path doesn't start with base_path, return as is
            return full_path
    
    
    # ____________________ FILE INFO ____________________ #
    def get_file_type(self, file_path: str) -> str:
        """Get the file extension from a file path"""
        return Path(file_path).suffix.lower()

    def get_file_info(self, file_path: str) -> Dict[str, Any]:
        """Get comprehensive information about a file"""
        try:
            target_path = self.translate_in(file_path)
            if not self._exists_raw(target_path):
                return {"success": False, "error": f"File {file_path} does not exist"}
            
            # For now, return basic info - subclasses can override for more details
            return {
                "success": True,
                "file_path": target_path,
                "file_name": Path(target_path).name,
                "file_extension": Path(target_path).suffix.lower(),
                "exists": True
            }
        except Exception as e:
            logger.error(f"Error getting file info for {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def create_directory(self, path: str) -> Dict[str, Any]:
        """Create directory"""
        try:
            target_path = self.translate_in(path)
            success = self._create_directory_raw(target_path)
            if success:
                return {"success": True, "path": target_path, "message": "Directory created successfully"}
            else:
                return {"success": False, "error": "Failed to create directory", "path": target_path}
        except Exception as e:
            logger.error(f"Error creating directory {path}: {str(e)}")
            return {"success": False, "error": str(e), "path": path}
    
    def exists(self, path: str) -> bool:
        """Check if path exists"""
        target_path = self.translate_in(path)
        return self._exists_raw(target_path)
    
    
    # ____________________ CURD ____________________ #
    def delete(self, path: str) -> Dict[str, Any]:
        """Delete file or directory"""
        try:
            target_path = self.translate_in(path)
            success = self._delete_raw(target_path)
            if success:
                return {"success": True, "path": target_path, "message": "Deleted successfully"}
            else:
                return {"success": False, "error": "Failed to delete", "path": target_path}
        except Exception as e:
            logger.error(f"Error deleting {path}: {str(e)}")
            return {"success": False, "error": str(e), "path": path}
    
    def move(self, source: str, destination: str) -> Dict[str, Any]:
        """Move/rename file or directory"""
        try:
            resolved_source = self.translate_in(source)
            resolved_destination = self.translate_in(destination)
            
            # Read source content
            content = self._read_raw(resolved_source)
            
            # Write to destination
            success = self._write_raw(resolved_destination, content)
            if success:
                # Delete source
                self._delete_raw(resolved_source)
                return {"success": True, "source": resolved_source, "destination": resolved_destination, "message": "Moved successfully"}
            else:
                return {"success": False, "error": "Failed to write to destination", "source": resolved_source, "destination": resolved_destination}
        except Exception as e:
            logger.error(f"Error moving {source} to {destination}: {str(e)}")
            return {"success": False, "error": str(e), "source": source, "destination": destination}
    
    def copy(self, source: str, destination: str) -> Dict[str, Any]:
        """Copy file"""
        try:
            resolved_source = self.translate_in(source)
            resolved_destination = self.translate_in(destination)
            
            # Read source content
            content = self._read_raw(resolved_source)
            
            # Write to destination
            success = self._write_raw(resolved_destination, content)
            if success:
                return {"success": True, "source": resolved_source, "destination": resolved_destination, "message": "Copied successfully"}
            else:
                return {"success": False, "error": "Failed to write to destination", "source": resolved_source, "destination": resolved_destination}
        except Exception as e:
            logger.error(f"Error copying {source} to {destination}: {str(e)}")
            return {"success": False, "error": str(e), "source": source, "destination": destination}
    
    def list(self, path: str = None, max_depth: int = 3, include_hidden: bool = False) -> Dict[str, Any]:
        """List files and directories"""
        try:
            target_path = self.translate_in(path) if path else str(self.base_path)
            items = self._list_raw(target_path, max_depth=max_depth, include_hidden=include_hidden)
            
            return {
                "success": True,
                "path": target_path,
                "items": items,
                "total_count": len(items)
            }
        except Exception as e:
            logger.error(f"Error listing {path}: {str(e)}")
            return {"success": False, "error": str(e), "path": path}
    
    def save(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """
        Save content to a file with automatic format detection.
        This method replaces the old save method with the improved create_file logic.
        
        Args:
            file_path (str): Path where the file should be saved
            content (Any): Content to save to the file
            **kwargs: Additional arguments for file creation (encoding, format, etc.)
            
        Returns:
            Dict[str, Any]: Result of the operation with success status and details
        """
        try:
            # Get file type to determine the appropriate save method
            file_extension = self.get_file_type(file_path)
            target_file_path = self.translate_in(file_path)
            
            # Route to specialized save methods based on file type
            if file_extension == '.json':
                return self._save_json(target_file_path, content, **kwargs)
            elif file_extension in ['.txt', '.md', '.log']:
                return self._save_text(target_file_path, content, **kwargs)
            elif file_extension == '.csv':
                return self._save_csv(target_file_path, content, **kwargs)
            elif file_extension in ['.yaml', '.yml']:
                return self._save_yaml(target_file_path, content, **kwargs)
            elif file_extension == '.xml':
                return self._save_xml(target_file_path, content, **kwargs)
            elif file_extension == '.xlsx':
                return self._save_excel(target_file_path, content, **kwargs)
            elif file_extension == '.pickle':
                return self._save_pickle(target_file_path, content, **kwargs)
            elif file_extension == '.pdf':
                return self._save_pdf(target_file_path, content, **kwargs)
            elif file_extension in ['.png', '.jpg', '.jpeg', '.gif', '.bmp', '.tiff']:
                return self._save_image(target_file_path, content, **kwargs)
            else:
                # For other file types, use the generic approach
                # Convert content to bytes if it's not already
                if isinstance(content, str):
                    content_bytes = content.encode(kwargs.get('encoding', 'utf-8'))
                elif isinstance(content, bytes):
                    content_bytes = content
                else:
                    content_bytes = str(content).encode(kwargs.get('encoding', 'utf-8'))
                
                # Write the file using the raw method
                success = self._write_raw(target_file_path, content_bytes, **kwargs)
            
                if success:
                    return {
                        "success": True,
                        "message": f"File '{file_path}' saved successfully",
                        "file_path": file_path,
                        "full_path": target_file_path,
                        "size": len(content_bytes)
                    }
                else:
                    return {
                        "success": False,
                        "message": f"Failed to save file '{file_path}'",
                        "file_path": file_path,
                        "full_path": target_file_path
                    }
        except Exception as e:
            logger.error(f"Error saving file {file_path}: {str(e)}")
            return {
                "success": False,
                "message": f"Error saving file: {str(e)}",
                "file_path": file_path
            }
    
    def read(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read content from a file with automatic format detection"""
        try:
            target_file_path = self.translate_in(file_path)
            file_extension = Path(target_file_path).suffix.lower()
            
            # Handle different file types
            if file_extension == '.json':
                return self._read_json(target_file_path, **kwargs)
            elif file_extension in ['.yaml', '.yml']:
                return self._read_yaml(target_file_path, **kwargs)
            elif file_extension == '.csv':
                return self._read_csv(target_file_path, **kwargs)
            elif file_extension == '.xlsx':
                return self._read_excel(target_file_path, **kwargs)
            elif file_extension == '.xml':
                return self._read_xml(target_file_path, **kwargs)
            elif file_extension == '.pickle':
                return self._read_pickle(target_file_path, **kwargs)
            elif file_extension == '.pdf':
                return self._read_pdf(target_file_path, **kwargs)
            elif file_extension in ['.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff']:
                return self._read_image(target_file_path, **kwargs)
            else:
                # Default to text
                return self._read_text(target_file_path, **kwargs)
                
        except Exception as e:
            logger.error(f"Error reading {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def append(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to a file (only for supported formats)"""
        try:
            target_file_path = self.translate_in(file_path)
            file_extension = Path(target_file_path).suffix.lower()
            
            if file_extension in self.appendable_formats:
                return self.appendable_formats[file_extension](target_file_path, content, **kwargs)
            else:
                return {"success": False, "error": f"Append not supported for {file_extension} files"}
                
        except Exception as e:
            logger.error(f"Error appending to {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Text file handlers
    def _save_text(self, file_path: str, content: Any, encoding: str = 'utf-8', **kwargs) -> Dict[str, Any]:
        """Save text content to a file"""
        try:
            # Convert content to bytes
            if isinstance(content, str):
                content_bytes = content.encode(encoding)
            else:
                content_bytes = str(content).encode(encoding)
            
            # Use raw write method
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"File saved to {file_path}",
                    "file_path": file_path,
                    "content_length": len(content_bytes)
                }
            else:
                return {"success": False, "error": "Failed to write file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error saving text file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_text(self, file_path: str, encoding: str = 'utf-8', **kwargs) -> Dict[str, Any]:
        """Read text content from a file"""
        try:
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            content = content_bytes.decode(encoding)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path,
                "content_length": len(content)
            }
        except Exception as e:
            logger.error(f"Error reading text file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
        
    def _append_text(self, file_path: str, content: str, encoding: str = 'utf-8', **kwargs) -> Dict[str, Any]:
        """Append text content to a file"""
        try:
            # Convert content to bytes
            content_bytes = str(content).encode(encoding)
            
            # For append, read existing content first
            existing_bytes = b""
            if self._exists_raw(file_path):
                existing_bytes = self._read_raw(file_path, **kwargs)
            
            # Combine existing and new content
            combined_bytes = existing_bytes + content_bytes
            
            # Write combined content
            success = self._write_raw(file_path, combined_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Content appended to file {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to append to file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error appending to text file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # JSON file handlers
    def _save_json(self, file_path: str, content: Any, indent: int = 2, **kwargs) -> Dict[str, Any]:
        """Save JSON content to a file"""
        try:
            # Convert content to JSON string
            if isinstance(content, str):
                # Validate JSON
                json.loads(content)
                json_content = content
            else:
                json_content = json.dumps(content, indent=indent, ensure_ascii=False)
            
            # Convert to bytes
            content_bytes = json_content.encode('utf-8')
            
            # Use raw write method
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"JSON file saved to {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to write file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error saving JSON file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_json(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read JSON content from a file"""
        try:
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            content_str = content_bytes.decode('utf-8')
            
            # Parse JSON
            content = json.loads(content_str)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading JSON file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_json(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to JSON file (for arrays)"""
        try:
            # Read existing content
            existing_content = []
            if self._exists_raw(file_path):
                existing_bytes = self._read_raw(file_path, **kwargs)
                existing_str = existing_bytes.decode('utf-8')
                existing_content = json.loads(existing_str)
            
            # Merge content
            if isinstance(existing_content, list):
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            elif isinstance(existing_content, dict):
                if isinstance(content, dict):
                    existing_content.update(content)
                else:
                    return {"success": False, "error": "Cannot append non-dict to JSON dict"}
            else:
                existing_content = [existing_content]
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            
            # Convert to JSON string and bytes
            json_content = json.dumps(existing_content, indent=2, ensure_ascii=False)
            content_bytes = json_content.encode('utf-8')
            
            # Write combined content
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Content appended to JSON file {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to append to file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error appending to JSON file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # CSV file handlers
    def _save_csv(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save CSV content to a file - handles both raw CSV strings and structured data"""
        try:
            if not content:
                return {"success": False, "error": "No content to save"}
            
            from io import StringIO
            
            # Build CSV content in memory
            csv_buffer = StringIO()
            
            # If content is a string, use it directly
            if isinstance(content, str):
                csv_content = content
                rows = content.count('\n')
            # If content is a list of dictionaries, use CSV writer
            elif isinstance(content, list) and content and isinstance(content[0], dict):
                fieldnames = content[0].keys()
                writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(content)
                csv_content = csv_buffer.getvalue()
                rows = len(content)
            # If content is a list of lists, use CSV writer
            elif isinstance(content, list) and content and isinstance(content[0], list):
                writer = csv.writer(csv_buffer)
                writer.writerows(content)
                csv_content = csv_buffer.getvalue()
                rows = len(content)
            else:
                return {"success": False, "error": "CSV content must be a string, list of dictionaries, or list of lists"}
            
            # Convert to bytes and write
            content_bytes = csv_content.encode('utf-8')
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"CSV file saved to {file_path}",
                    "file_path": file_path,
                    "rows": rows
                }
            else:
                return {"success": False, "error": "Failed to write file", "file_path": file_path}
            
        except Exception as e:
            logger.error(f"Error saving CSV file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_csv(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read CSV content from a file"""
        try:
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            content_str = content_bytes.decode('utf-8')
            
            # Parse CSV
            from io import StringIO
            reader = csv.DictReader(StringIO(content_str))
            content = list(reader)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path,
                "rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error reading CSV file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_csv(self, file_path: str, content: List[Dict[str, Any]], **kwargs) -> Dict[str, Any]:
        """Append content to CSV file"""
        try:
            if not content:
                return {"success": False, "error": "No content to append"}
            
            # Read existing content
            existing_content = []
            if self._exists_raw(file_path):
                existing_bytes = self._read_raw(file_path, **kwargs)
                existing_str = existing_bytes.decode('utf-8')
                from io import StringIO
                reader = csv.DictReader(StringIO(existing_str))
                existing_content = list(reader)
            
            # Combine content
            combined_content = existing_content + content
            
            # Write combined content
            from io import StringIO
            csv_buffer = StringIO()
            if combined_content:
                fieldnames = combined_content[0].keys()
                writer = csv.DictWriter(csv_buffer, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(combined_content)
            
            csv_content = csv_buffer.getvalue()
            content_bytes = csv_content.encode('utf-8')
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Content appended to CSV file {file_path}",
                    "file_path": file_path,
                    "appended_rows": len(content)
                }
            else:
                return {"success": False, "error": "Failed to append to file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error appending to CSV file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # YAML file handlers
    def _save_yaml(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save YAML content to a file"""
        try:
            # Convert content to YAML string
            yaml_content = yaml.dump(content, default_flow_style=False, allow_unicode=True)
            content_bytes = yaml_content.encode('utf-8')
            
            # Use raw write method
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"YAML file saved to {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to write file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error saving YAML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_yaml(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read YAML content from a file"""
        try:
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            content_str = content_bytes.decode('utf-8')
            
            # Parse YAML
            content = yaml.safe_load(content_str)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading YAML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_yaml(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to YAML file (for lists)"""
        try:
            # Read existing content
            existing_content = []
            if self._exists_raw(file_path):
                existing_bytes = self._read_raw(file_path, **kwargs)
                existing_str = existing_bytes.decode('utf-8')
                existing_content = yaml.safe_load(existing_str) or []
            
            # Merge content
            if isinstance(existing_content, list):
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            elif isinstance(existing_content, dict):
                if isinstance(content, dict):
                    existing_content.update(content)
                else:
                    return {"success": False, "error": "Cannot append non-dict to YAML dict"}
            else:
                existing_content = [existing_content]
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            
            # Convert to YAML string and bytes
            yaml_content = yaml.dump(existing_content, default_flow_style=False, allow_unicode=True)
            content_bytes = yaml_content.encode('utf-8')
            
            # Write combined content
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Content appended to YAML file {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to append to file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error appending to YAML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # XML file handlers
    def _save_xml(self, file_path: str, content: Any, root_tag: str = "root", **kwargs) -> Dict[str, Any]:
        """Save XML content to a file"""
        try:
            # If content is already a string (raw XML), use it directly
            if isinstance(content, str):
                # Check if it's already valid XML
                try:
                    ET.fromstring(content)
                    xml_content = content
                except ET.ParseError:
                    # Not valid XML, treat as text content and wrap it
                    root = ET.Element(root_tag)
                    root.text = content
                    xml_content = ET.tostring(root, encoding='unicode')
            # If content is a dictionary, convert to XML
            elif isinstance(content, dict):
                def dict_to_xml(data, root):
                    for key, value in data.items():
                        child = ET.SubElement(root, key)
                        if isinstance(value, dict):
                            dict_to_xml(value, child)
                        else:
                            child.text = str(value)
                
                root = ET.Element(root_tag)
                dict_to_xml(content, root)
                xml_content = ET.tostring(root, encoding='unicode')
            else:
                # For other types, wrap in root element
                root = ET.Element(root_tag)
                root.text = str(content)
                xml_content = ET.tostring(root, encoding='unicode')
            
            # Convert to bytes and write
            content_bytes = xml_content.encode('utf-8')
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"XML file saved to {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to write file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error saving XML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_xml(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read XML content from a file"""
        try:
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            content_str = content_bytes.decode('utf-8')
            
            # Parse XML
            root = ET.fromstring(content_str)
            
            def xml_to_dict(element):
                result = {}
                for child in element:
                    if len(child) == 0:
                        result[child.tag] = child.text
                    else:
                        result[child.tag] = xml_to_dict(child)
                return result
            
            content = xml_to_dict(root)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading XML file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Excel file handlers
    def _save_excel(self, file_path: str, content: List[List[Any]], sheet_name: str = "Sheet1", **kwargs) -> Dict[str, Any]:
        """Save Excel content to a file"""
        if not EXCEL_AVAILABLE:
            return {"success": False, "error": "openpyxl library not available"}
        
        try:
            from io import BytesIO
            
            # Create workbook in memory
            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = sheet_name
            
            for row in content:
                worksheet.append(row)
            
            # Save to BytesIO buffer
            buffer = BytesIO()
            workbook.save(buffer)
            content_bytes = buffer.getvalue()
            
            # Use raw write method
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Excel file saved to {file_path}",
                    "file_path": file_path,
                    "rows": len(content)
                }
            else:
                return {"success": False, "error": "Failed to write file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error saving Excel file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_excel(self, file_path: str, sheet_name: str = None, **kwargs) -> Dict[str, Any]:
        """Read Excel content from a file"""
        if not EXCEL_AVAILABLE:
            return {"success": False, "error": "openpyxl library not available"}
        
        try:
            from io import BytesIO
            
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            
            # Load workbook from bytes
            workbook = load_workbook(BytesIO(content_bytes), data_only=True)
            sheet_names = workbook.sheetnames
            
            if sheet_name is None:
                sheet_name = sheet_names[0]
            
            if sheet_name not in sheet_names:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            worksheet = workbook[sheet_name]
            content = []
            
            for row in worksheet.iter_rows(values_only=True):
                if any(cell is not None for cell in row):
                    content.append(list(row))
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "rows": len(content)
            }
        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_excel(self, file_path: str, content: List[List[Any]], sheet_name: str = None, **kwargs) -> Dict[str, Any]:
        """Append content to Excel file"""
        if not EXCEL_AVAILABLE:
            return {"success": False, "error": "openpyxl library not available"}
        
        try:
            from io import BytesIO
            
            if not self._exists_raw(file_path):
                return self._save_excel(file_path, content, sheet_name or "Sheet1", **kwargs)
            
            # Read existing content
            content_bytes = self._read_raw(file_path, **kwargs)
            workbook = load_workbook(BytesIO(content_bytes))
            sheet_names = workbook.sheetnames
            
            if sheet_name is None:
                sheet_name = sheet_names[0]
            
            if sheet_name not in sheet_names:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            worksheet = workbook[sheet_name]
            
            for row in content:
                worksheet.append(row)
            
            # Save to BytesIO buffer
            buffer = BytesIO()
            workbook.save(buffer)
            updated_bytes = buffer.getvalue()
            
            # Write updated content
            success = self._write_raw(file_path, updated_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Content appended to Excel file {file_path}",
                    "file_path": file_path,
                    "appended_rows": len(content)
                }
            else:
                return {"success": False, "error": "Failed to append to file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error appending to Excel file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Pickle file handlers
    def _save_pickle(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save pickle content to a file"""
        try:
            # Convert content to bytes using pickle
            content_bytes = pickle.dumps(content)
            
            # Use raw write method
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Pickle file saved to {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to write file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error saving pickle file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_pickle(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read pickle content from a file"""
        try:
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            
            # Parse pickle content
            content = pickle.loads(content_bytes)
            
            return {
                "success": True,
                "content": content,
                "file_path": file_path
            }
        except Exception as e:
            logger.error(f"Error reading pickle file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _append_pickle(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Append content to pickle file (for lists)"""
        try:
            # Read existing content
            existing_content = []
            if self._exists_raw(file_path):
                existing_bytes = self._read_raw(file_path, **kwargs)
                existing_content = pickle.loads(existing_bytes)
            
            # Merge content
            if isinstance(existing_content, list):
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            elif isinstance(existing_content, dict):
                if isinstance(content, dict):
                    existing_content.update(content)
                elif isinstance(content, list):
                    existing_content["appended_list"] = content
                else:
                    existing_content["appended_value"] = content
            else:
                existing_content = [existing_content]
                if isinstance(content, list):
                    existing_content.extend(content)
                else:
                    existing_content.append(content)
            
            # Convert to bytes and write
            content_bytes = pickle.dumps(existing_content)
            success = self._write_raw(file_path, content_bytes, **kwargs)
            
            if success:
                return {
                    "success": True,
                    "message": f"Content appended to pickle file {file_path}",
                    "file_path": file_path
                }
            else:
                return {"success": False, "error": "Failed to append to file", "file_path": file_path}
        except Exception as e:
            logger.error(f"Error appending to pickle file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # PDF file handlers
    def _save_pdf(self, file_path: str, content: str, **kwargs) -> Dict[str, Any]:
        """Save content to a PDF file"""
        try:
            # Use reportlab to create a proper PDF with text content
            from reportlab.lib.pagesizes import letter
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
            
            # Create PDF document
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            styles = getSampleStyleSheet()
            story = []
            
            # Split content into paragraphs
            paragraphs = content.split('\n')
            
            for para_text in paragraphs:
                if para_text.strip():  # Non-empty paragraph
                    para = Paragraph(para_text, styles['Normal'])
                    story.append(para)
                    story.append(Spacer(1, 12))
                else:
                    story.append(Spacer(1, 12))
            
            # Build the PDF
            doc.build(story)
            
            return {
                "success": True,
                "message": f"PDF file saved to {file_path}",
                "file_path": file_path
            }
                
        except ImportError:
            return {"success": False, "error": "reportlab library not available for PDF creation"}
        except Exception as e:
            logger.error(f"Error saving PDF file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_pdf(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read content from a PDF file"""
        if not PDF_AVAILABLE:
            return {"success": False, "error": "unstructured library not available"}
        try:
            doc = pymupdf.open(file_path)
            all_text = []
            for page in doc:
                text = page.get_text()
                all_text.append(text)
            text = "\n\n".join(all_text)
            return {
                "success": True,
                "content": text,
                "file_path": file_path
            }
                
        except Exception as e:
            logger.error(f"Error reading PDF file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Image file handlers
    def _save_image(self, file_path: str, content: Any, **kwargs) -> Dict[str, Any]:
        """Save image content to a file"""
        if not PILLOW_AVAILABLE:
            return {"success": False, "error": "Pillow library not available"}
        
        try:
            from io import BytesIO
            
            # Check if content is already a PIL Image object
            if hasattr(content, 'save') and callable(getattr(content, 'save', None)):
                # Content is a PIL Image object - save to BytesIO first
                buffer = BytesIO()
                content.save(buffer, format=content.format or 'PNG')
                content_bytes = buffer.getvalue()
                
                # Use raw write method
                success = self._write_raw(file_path, content_bytes, **kwargs)
                
                if success:
                    return {
                        "success": True,
                        "message": f"Image saved to {file_path}",
                        "file_path": file_path,
                        "format": content.format,
                        "size": content.size
                    }
                else:
                    return {"success": False, "error": "Failed to write file", "file_path": file_path}
            elif isinstance(content, bytes):
                # Content is binary image data
                success = self._write_raw(file_path, content, **kwargs)
                
                if success:
                    return {
                        "success": True,
                        "message": f"Image saved to {file_path}",
                        "file_path": file_path
                    }
                else:
                    return {"success": False, "error": "Failed to write file", "file_path": file_path}
            elif isinstance(content, str) and Path(content).exists():
                # Content is a file path to an existing image - read and write
                with open(content, 'rb') as f:
                    content_bytes = f.read()
                
                success = self._write_raw(file_path, content_bytes, **kwargs)
                
                if success:
                    return {
                        "success": True,
                        "message": f"Image copied from {content} to {file_path}",
                        "file_path": file_path
                    }
                else:
                    return {"success": False, "error": "Failed to write file", "file_path": file_path}
            else:
                return {"success": False, "error": "Content must be a PIL Image object, binary data, or valid file path"}
                
        except Exception as e:
            logger.error(f"Error saving image file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    def _read_image(self, file_path: str, **kwargs) -> Dict[str, Any]:
        """Read image and return PIL Image object"""
        if not PILLOW_AVAILABLE:
            return {"success": False, "error": "Pillow library not available"}
        
        try:
            from io import BytesIO
            
            # Use raw read method
            content_bytes = self._read_raw(file_path, **kwargs)
            
            # Open image from bytes
            with Image.open(BytesIO(content_bytes)) as img:
                # Convert to RGB if necessary for consistency
                if img.mode in ('RGBA', 'LA', 'P'):
                    img = img.convert('RGB')
                
                metadata = {
                    "format": img.format,
                    "mode": img.mode,
                    "size": img.size,
                    "width": img.width,
                    "height": img.height
                }
                
                return {
                    "success": True,
                    "content": img,  # Return the PIL Image object
                    "metadata": metadata,
                    "file_path": file_path
                }
                
        except Exception as e:
            logger.error(f"Error reading image file {file_path}: {str(e)}")
            return {"success": False, "error": str(e), "file_path": file_path}
    
    # Placeholder for future database integration
    def _get_database_connection(self, db_type: str, connection_string: str) -> Any:
        """Placeholder for future database integration"""
        # This will be implemented when adding database support
        raise NotImplementedError("Database integration not yet implemented") 